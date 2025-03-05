import csv
import io

from pypdf import PdfReader
from zipfile import ZipFile
from tests.conftest import ARCHIVE_FILE_PATH
from openpyxl import load_workbook


def test_pdf():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("program.pdf") as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            assert len(pdf_reader.pages) > 0
            assert 'Работаем с файлами. Александр Котляр' in pdf_reader.pages[1].extract_text()


def test_csv():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("students.csv") as csv_file:
            csv_content = csv_file.read().decode('utf-8')
            items_list = list(csv.reader(csv_content.splitlines()))
            assert len(items_list) == 3
            assert items_list[0][3] == '88'


def test_xlsx():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("report.xlsx") as xlsx_file:
            file_content = io.BytesIO(xlsx_file.read())
            workbook = load_workbook(file_content)
            workbook_content = workbook.active
            assert workbook_content.max_row == 4
            assert workbook_content.max_column == 5
            assert workbook_content['E2'].value == '=AVERAGE(B2:D2)'